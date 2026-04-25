from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = (
    ROOT
    / "pipeline"
    / "workbooks"
    / "releases"
    / "recipe-menu-costing-workbook-v1.xlsx"
)


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D0D7DE")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row: int, headers: list[str]) -> None:
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=value)
        cell.fill = TITLE_FILL
        cell.font = WHITE_FONT
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_section(cell) -> None:
    cell.fill = SECTION_FILL
    cell.font = BOLD
    cell.border = BOX


def style_input(cell) -> None:
    cell.fill = INPUT_FILL
    cell.border = BOX


def style_formula(cell) -> None:
    cell.fill = FORMULA_FILL
    cell.border = BOX
    cell.protection = Protection(locked=True)


def style_text(cell) -> None:
    cell.border = BOX
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def fit_columns(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_note(ws, row: int, title: str, body: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=title)
    style_section(title_cell)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 2, end_column=6)
    body_cell = ws.cell(row=row + 1, column=1, value=body)
    body_cell.alignment = Alignment(wrap_text=True, vertical="top")
    body_cell.border = BOX
    return row + 4


def build_start_here(ws) -> None:
    ws.title = "Start Here"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Recipe and Menu Costing Workbook"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = "Cost dishes clearly, price menus confidently, and update margins faster when ingredient costs change."
    ws["A2"].alignment = Alignment(wrap_text=True)

    row = 4
    row = add_note(
        ws,
        row,
        "How to use this workbook",
        "1. Set your defaults in Settings.\n"
        "2. Add ingredients in Ingredient Master.\n"
        "3. Build prep items in Sub-Recipes if needed.\n"
        "4. Cost dishes in Recipe Costing.\n"
        "5. Review suggested prices in Menu Pricing.\n"
        "6. Check the Summary Dashboard for weak margins.",
    )
    row = add_note(
        ws,
        row,
        "Color guide",
        "Yellow cells are for inputs.\n"
        "Green cells contain formulas and should usually not be edited.\n"
        "Orange-tinted notes call out risks or assumptions.",
    )
    row = add_note(
        ws,
        row,
        "Best first step",
        "Start with one recipe only. Once the ingredient list and unit conversions look right, duplicate the structure for the rest of your menu.",
    )

    ws["A17"] = "Common terms"
    style_section(ws["A17"])
    terms = [
        ("Food cost %", "The share of the selling price taken up by food cost."),
        ("Purchase unit", "How you buy the ingredient, such as kg, case, bottle, or pack."),
        ("Recipe unit", "How the ingredient is used in recipes, such as g, ml, or each."),
        ("Yield %", "The usable share of the purchased item after trim, peel, or waste."),
    ]
    r = 18
    for term, desc in terms:
        ws[f"A{r}"] = term
        ws[f"A{r}"].font = BOLD
        ws[f"B{r}"] = desc
        style_text(ws[f"A{r}"])
        style_text(ws[f"B{r}"])
        r += 1
    fit_columns(ws, {"A": 22, "B": 70, "C": 14, "D": 14, "E": 14, "F": 14})


def build_settings(ws) -> None:
    ws["A1"] = "Settings"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Business Defaults"
    style_section(ws["A3"])
    labels = [
        ("Currency Symbol", "$"),
        ("Default Target Food Cost %", 0.30),
        ("Default Gross Margin %", 0.70),
        ("Default Yield %", 1.00),
    ]
    row = 4
    for label, value in labels:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        style_text(ws[f"A{row}"])
        style_input(ws[f"B{row}"])
        row += 1

    ws["D3"] = "Supported Units"
    style_section(ws["D3"])
    units = ["kg", "g", "lb", "oz", "l", "ml", "each", "pack", "bottle", "case"]
    for idx, unit in enumerate(units, start=4):
        ws[f"D{idx}"] = unit
        style_text(ws[f"D{idx}"])

    ws["F3"] = "Categories"
    style_section(ws["F3"])
    categories = ["Protein", "Produce", "Dry Goods", "Dairy", "Bakery", "Beverage", "Other"]
    for idx, category in enumerate(categories, start=4):
        ws[f"F{idx}"] = category
        style_text(ws[f"F{idx}"])

    fit_columns(ws, {"A": 28, "B": 18, "C": 4, "D": 18, "E": 4, "F": 18})


def build_ingredient_master(ws) -> None:
    headers = [
        "Ingredient Name",
        "Category",
        "Purchase Unit",
        "Purchase Price",
        "Recipe Unit",
        "Units per Purchase Unit",
        "Yield %",
        "Cost per Recipe Unit",
        "Supplier Notes",
    ]
    style_header(ws, 1, headers)
    for row in range(2, 202):
        for col in range(1, len(headers) + 1):
            style_text(ws.cell(row=row, column=col))
        for col in [1, 2, 3, 4, 5, 6, 7, 9]:
            style_input(ws.cell(row=row, column=col))
        formula = f'=IF(OR(D{row}="",F{row}="",G{row}=""),"",IFERROR(D{row}/(F{row}*G{row}),"CHECK INPUT"))'
        ws[f"H{row}"] = formula
        style_formula(ws[f"H{row}"])

    ws.freeze_panes = "A2"
    fit_columns(
        ws,
        {"A": 26, "B": 16, "C": 16, "D": 14, "E": 14, "F": 20, "G": 10, "H": 18, "I": 24},
    )


def build_subrecipes(ws) -> None:
    headers = [
        "Sub-Recipe Name",
        "Batch Yield",
        "Ingredient Name",
        "Qty Used",
        "Unit Cost",
        "Line Cost",
        "Batch Cost",
        "Cost per Yield Unit",
    ]
    style_header(ws, 1, headers)
    for row in range(2, 202):
        for col in range(1, len(headers) + 1):
            style_text(ws.cell(row=row, column=col))
        for col in [1, 2, 3, 4]:
            style_input(ws.cell(row=row, column=col))
        ws[f"E{row}"] = f'=IF(C{row}="","",IFERROR(XLOOKUP(C{row},\'Ingredient Master\'!$A$2:$A$201,\'Ingredient Master\'!$H$2:$H$201,""),""))'
        ws[f"F{row}"] = f'=IF(OR(D{row}="",E{row}=""),"",D{row}*E{row})'
        ws[f"G{row}"] = f'=IF(A{row}="","",SUMIF($A$2:$A$201,A{row},$F$2:$F$201))'
        ws[f"H{row}"] = f'=IF(OR(B{row}="",G{row}=""),"",G{row}/B{row})'
        for col in ["E", "F", "G", "H"]:
            style_formula(ws[f"{col}{row}"])

    ws.freeze_panes = "A2"
    fit_columns(ws, {"A": 24, "B": 12, "C": 24, "D": 12, "E": 12, "F": 12, "G": 12, "H": 16})


def build_recipe_costing(ws) -> None:
    headers = [
        "Recipe Name",
        "Serving Count",
        "Line Type",
        "Item Name",
        "Qty Used",
        "Unit Cost",
        "Line Cost",
        "Batch Cost",
        "Portion Cost",
    ]
    style_header(ws, 1, headers)
    for row in range(2, 302):
        for col in range(1, len(headers) + 1):
            style_text(ws.cell(row=row, column=col))
        for col in [1, 2, 3, 4, 5]:
            style_input(ws.cell(row=row, column=col))

        ws[f"F{row}"] = (
            f'=IF(D{row}="","",IF(C{row}="Ingredient",'
            f'IFERROR(XLOOKUP(D{row},\'Ingredient Master\'!$A$2:$A$201,\'Ingredient Master\'!$H$2:$H$201,""),""),'
            f'IF(C{row}="Sub-Recipe",IFERROR(XLOOKUP(D{row},\'Sub-Recipes\'!$A$2:$A$201,\'Sub-Recipes\'!$H$2:$H$201,""),""),"")))'
        )
        ws[f"G{row}"] = f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})'
        ws[f"H{row}"] = f'=IF(A{row}="","",SUMIF($A$2:$A$301,A{row},$G$2:$G$301))'
        ws[f"I{row}"] = f'=IF(OR(B{row}="",H{row}=""),"",H{row}/B{row})'
        for col in ["F", "G", "H", "I"]:
            style_formula(ws[f"{col}{row}"])

    ws.freeze_panes = "A2"
    fit_columns(ws, {"A": 24, "B": 12, "C": 14, "D": 24, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12})


def build_menu_pricing(ws) -> None:
    headers = [
        "Recipe Name",
        "Portion Cost",
        "Target Food Cost %",
        "Effective Food Cost %",
        "Suggested Menu Price",
        "Current Price",
        "Current Food Cost %",
        "Price Gap",
    ]
    style_header(ws, 1, headers)
    for row in range(2, 102):
        for col in range(1, len(headers) + 1):
            style_text(ws.cell(row=row, column=col))
        for col in ["A", "C", "F"]:
            style_input(ws[f"{col}{row}"])
        ws[f"B{row}"] = f'=IF(A{row}="","",IFERROR(XLOOKUP(A{row},\'Recipe Costing\'!$A$2:$A$301,\'Recipe Costing\'!$I$2:$I$301,""),""))'
        ws[f"D{row}"] = f'=IF(C{row}="",Settings!$B$5,C{row})'
        ws[f"E{row}"] = f'=IF(OR(B{row}="",D{row}=""),"",IFERROR(B{row}/D{row},""))'
        ws[f"G{row}"] = f'=IF(OR(B{row}="",F{row}=""),"",IFERROR(B{row}/F{row},""))'
        ws[f"H{row}"] = f'=IF(OR(E{row}="",F{row}=""),"",F{row}-E{row})'
        for col in ["B", "D", "E", "G", "H"]:
            style_formula(ws[f"{col}{row}"])

    ws.freeze_panes = "A2"
    fit_columns(ws, {"A": 24, "B": 12, "C": 16, "D": 18, "E": 18, "F": 14, "G": 16, "H": 12})


def build_summary(ws) -> None:
    ws["A1"] = "Summary Dashboard"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Review this page after updating ingredient prices or recipe servings."
    ws["A5"] = "Key defaults"
    style_section(ws["A5"])
    ws["A6"] = "Default target food cost %"
    ws["B6"] = "=Settings!B5"
    style_text(ws["A6"])
    style_formula(ws["B6"])

    headers = ["Recipe Name", "Portion Cost", "Suggested Price", "Current Price", "Current Food Cost %", "Price Gap"]
    style_header(ws, 9, headers)
    for row in range(10, 60):
        for col in range(1, len(headers) + 1):
            style_text(ws.cell(row=row, column=col))
        ws[f"A{row}"] = f"=IF('Menu Pricing'!A{row-8}=\"\",\"\",'Menu Pricing'!A{row-8})"
        ws[f"B{row}"] = f"=IF(A{row}=\"\",\"\",'Menu Pricing'!B{row-8})"
        ws[f"C{row}"] = f"=IF(A{row}=\"\",\"\",'Menu Pricing'!E{row-8})"
        ws[f"D{row}"] = f"=IF(A{row}=\"\",\"\",'Menu Pricing'!F{row-8})"
        ws[f"E{row}"] = f"=IF(A{row}=\"\",\"\",'Menu Pricing'!G{row-8})"
        ws[f"F{row}"] = f"=IF(A{row}=\"\",\"\",'Menu Pricing'!H{row-8})"
        for col in ["A", "B", "C", "D", "E", "F"]:
            style_formula(ws[f"{col}{row}"])

    ws["H5"] = "Review prompts"
    style_section(ws["H5"])
    prompts = [
        "Recipes with a high positive price gap may be underpriced.",
        "Check any recipe where current food cost % is above your target.",
        "If many recipes change when one ingredient price moves, that is a strong app-upgrade signal.",
    ]
    for idx, prompt in enumerate(prompts, start=6):
        ws[f"H{idx}"] = prompt
        ws[f"H{idx}"].fill = WARN_FILL
        ws[f"H{idx}"].border = BOX
        ws[f"H{idx}"].alignment = Alignment(wrap_text=True)

    fit_columns(ws, {"A": 24, "B": 12, "C": 16, "D": 14, "E": 18, "F": 12, "H": 42})


def add_validations(wb: Workbook) -> None:
    settings = wb["Settings"]
    ingredient = wb["Ingredient Master"]
    subrecipes = wb["Sub-Recipes"]
    recipes = wb["Recipe Costing"]
    pricing = wb["Menu Pricing"]

    unit_validation = DataValidation(
        type="list",
        formula1="=Settings!$D$4:$D$13",
        allow_blank=True,
    )
    category_validation = DataValidation(
        type="list",
        formula1="=Settings!$F$4:$F$10",
        allow_blank=True,
    )
    line_type_validation = DataValidation(
        type="list",
        formula1='"Ingredient,Sub-Recipe"',
        allow_blank=True,
    )
    ingredient_name_validation = DataValidation(
        type="list",
        formula1="='Ingredient Master'!$A$2:$A$201",
        allow_blank=True,
    )
    subrecipe_name_validation = DataValidation(
        type="list",
        formula1="='Sub-Recipes'!$A$2:$A$201",
        allow_blank=True,
    )
    recipe_name_validation = DataValidation(
        type="list",
        formula1="='Recipe Costing'!$A$2:$A$301",
        allow_blank=True,
    )

    ingredient.add_data_validation(unit_validation)
    ingredient.add_data_validation(category_validation)
    ingredient.add_data_validation(ingredient_name_validation)
    ingredient.add_data_validation(subrecipe_name_validation)
    subrecipes.add_data_validation(ingredient_name_validation)
    recipes.add_data_validation(line_type_validation)
    pricing.add_data_validation(recipe_name_validation)

    category_validation.add("B2:B201")
    unit_validation.add("C2:C201")
    unit_validation.add("E2:E201")
    ingredient_name_validation.add("C2:C201")
    line_type_validation.add("C2:C301")
    recipe_name_validation.add("A2:A101")


def protect_formula_sheets(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = "foodcosting"
        for row in ws.iter_rows():
            for cell in row:
                fill_color = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                if fill_color == INPUT_FILL.fgColor.rgb:
                    cell.protection = Protection(locked=False)


def add_sheet_tips(wb: Workbook) -> None:
    tips = {
        "Ingredient Master": "Add each ingredient once. Cost per recipe unit updates automatically from purchase price, units per purchase unit, and yield.",
        "Sub-Recipes": "Use this for sauces, doughs, batters, or prep items reused across several dishes.",
        "Recipe Costing": "Choose Ingredient or Sub-Recipe in Line Type, then select the item and quantity used.",
        "Menu Pricing": "Leave Target Food Cost % blank to use the default from Settings.",
    }
    for name, text in tips.items():
        ws = wb[name]
        ws["K1"] = "Tip"
        ws["K1"].fill = SECTION_FILL
        ws["K1"].font = BOLD
        ws["K2"] = text
        ws["K2"].fill = WARN_FILL
        ws["K2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["K2"].border = BOX
        ws.column_dimensions["K"].width = 36


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_start_here(wb.active)
    build_settings(wb.create_sheet("Settings"))
    build_ingredient_master(wb.create_sheet("Ingredient Master"))
    build_subrecipes(wb.create_sheet("Sub-Recipes"))
    build_recipe_costing(wb.create_sheet("Recipe Costing"))
    build_menu_pricing(wb.create_sheet("Menu Pricing"))
    build_summary(wb.create_sheet("Summary Dashboard"))

    add_validations(wb)
    add_sheet_tips(wb)
    protect_formula_sheets(wb)

    wb.save(OUTPUT)
    print(f"Wrote workbook to {OUTPUT}")


if __name__ == "__main__":
    main()
